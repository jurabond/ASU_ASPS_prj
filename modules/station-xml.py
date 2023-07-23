import mysql.connector
import tkinter as tk
from tkinter import ttk
from tkinter import Toplevel, Button, messagebox
from modules.config import db_config
import os
import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font, Color

def get_station_data(station_id):
    try:
        cnx = mysql.connector.connect(**db_config)
        cursor = cnx.cursor(dictionary=True)

        query = "SELECT * FROM stations WHERE ID = %s"
        cursor.execute(query, (station_id,))

        row = cursor.fetchone()

        cursor.close()
        cnx.close()

        return row if row else None

    except mysql.connector.Error as err:
        print(f"Something went wrong: {err}")
        return None

def update_station_activity(station_id, activity):
    try:
        cnx = mysql.connector.connect(**db_config)
        cursor = cnx.cursor(dictionary=True)

        query = "UPDATE stations SET Activity = %s WHERE ID = %s"
        cursor.execute(query, (activity, station_id))
        cnx.commit()

        # Реєстрація події у таблицю station_events
        event_query = "INSERT INTO station_events (station_id, event) VALUES (%s, %s)"
        cursor.execute(event_query, (station_id, activity))
        cnx.commit()

        cursor.close()
        cnx.close()

        return True

    except mysql.connector.Error as err:
        print(f"Something went wrong: {err}")
        return False

def show_station_info(station_id, x, y):
    data = get_station_data(station_id)
    if data:
        top = Toplevel()
        top.geometry(f"+{x}+{y}")  # Позиціонуємо вікно
        top.title(f"Дані про підстанцію {station_id}")

        # Базове повідомлення, що включає активність і температуру
        message = f"Активність: {data['Activity']}\n" \
                  f"Температура: {data['Temperature']}°C\n"

        # Якщо станція активна, додайте строчку про навантаження
        if data['Activity'] == 'Активна':
            message += f"Навантаження: {data['Load']}%\n"

        # Додати інформацію про час роботи до повідомлення
        message += f"Час роботи: {data['WorkingHours']} годин"

        # Відобразити повідомлення у вікні
        msg = tk.Label(top, text=message, font=("Helvetica", 16), bg="black", fg="white", justify="left")
        msg.pack()

        # Створення нового стилю
        style = ttk.Style()
        style.configure("TButton",
                        foreground="black",
                        background="white",
                        font=("Arial", 12),
                        padding=10)

        # Додати кнопки для управління станцією
        disable_button = ttk.Button(top, text="Вимкнути станцію", command=lambda: disable_station(station_id), style="TButton")
        disable_button.pack(pady=10)

        enable_button = ttk.Button(top, text="Увімкнути станцію", command=lambda: enable_station(station_id), style="TButton")
        enable_button.pack(pady=10)
    else:
        # Якщо не вдається отримати дані про станцію, відобразити повідомлення про помилку
        messagebox.showerror("Помилка", f"Не вдалося отримати дані про підстанцію {station_id}")


def disable_station(station_id):
    if update_station_activity(station_id, "Неактивна"):
        messagebox.showinfo("Станція вимкнена", f"Підстанція {station_id} вимкнена успішно!")
    else:
        messagebox.showerror("Помилка", f"Не вдалося вимкнути підстанцію {station_id}")

def enable_station(station_id):
    if update_station_activity(station_id, "Активна"):
        messagebox.showinfo("Станція увімкнена", f"Підстанція {station_id} увімкнена успішно!")
    else:
        messagebox.showerror("Помилка", f"Не вдалося увімкнути підстанцію {station_id}")

def update_station_values(station_id, temperature, load, voltage):
    try:
        cnx = mysql.connector.connect(**db_config)
        cursor = cnx.cursor(dictionary=True)

        query = "UPDATE stations SET Temperature = %s, Load = %s, Voltage = %s WHERE ID = %s"
        cursor.execute(query, (str(temperature), str(load), str(voltage), str(station_id)))
        cnx.commit()

        cursor.close()
        cnx.close()

        return True

    except mysql.connector.Error as err:
        print(f"Something went wrong: {err}")
        return False


def update_station_voltage(station_id, voltage):
    try:
        cnx = mysql.connector.connect(**db_config)
        cursor = cnx.cursor(dictionary=True)

        query = "UPDATE stations SET Voltage = %s WHERE ID = %s"
        cursor.execute(query, (str(voltage), str(station_id)))
        cnx.commit()

        cursor.close()
        cnx.close()

        return True

    except mysql.connector.Error as err:
        print(f"Something went wrong: {err}")
        return False

from openpyxl.styles import PatternFill

from openpyxl.styles import PatternFill

def turn_off_station(station_id):
    # Виконайте запит до бази даних для вимкнення станції
    success = update_station_activity(station_id, "Неактивна")
    if success:
        messagebox.showinfo("Станція вимкнена", f"Підстанція {station_id} вимкнена успішно!")

        # Створення нової книги Excel і вибір активного робочого аркуша
        wb = Workbook()
        ws = wb.active

        # Назви підстанцій
        station_names = ["Підстанція 1", "Підстанція 2", "Підстанція 3", "Підстанція 4"]
        ws.append([""] + station_names)

        # Визначення першої підстанції, яка була відключена
        first_disabled_station = (station_id - 1) % 4

        # Заповнення таблиці з графіком вимкнень
        for i in range(24):  # 24 години
            row = [f"{i:02d}:00"]
            for j in range(4):  # 4 підстанції
                cell = ws.cell(row=i + 2, column=j + 2)
                if j == first_disabled_station:
                    cell.fill = PatternFill(start_color="FF0000",
                                            end_color="FF0000",
                                            fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFFFFF",
                                            end_color="FFFFFF",
                                            fill_type="solid")
                if (i + 1) % 4 == 0:
                    cell.font = Font(color="FF0000")
                first_disabled_station = (first_disabled_station + 1) % 4  # Перехід до наступної підстанції
            ws.append(row)

        # Збереження файлу
        wb.save('outage_schedule.xlsx')
    else:
        messagebox.showerror("Помилка", f"Не вдалося вимкнути підстанцію {station_id}")